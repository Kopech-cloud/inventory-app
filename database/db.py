import json
import pymysql
from pymysql.cursors import DictCursor
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import bcrypt

import os
from dotenv import load_dotenv

load_dotenv()

DB_CONFIG = {
    "host": os.getenv("DB_HOST"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD"),
    "database": os.getenv("DB_NAME"),
    "port": int(os.getenv("DB_PORT", 3306)),
    "cursorclass": pymysql.cursors.DictCursor
}


def get_connection():
    return pymysql.connect(**DB_CONFIG)


def init_db():
    from database.schema import create_tables
    create_tables()


def create_default_admin():
    from database.seed import seed_default_admin
    seed_default_admin()


def calculate_status(stock_qty: int) -> str:
    if stock_qty <= 0:
        return "Out of Stock"
    elif stock_qty <= 5:
        return "Low Stock"
    return "In Stock"


# ---------------- USERS ----------------

def authenticate_user(username: str, password: str):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, username, password, role
        FROM users
        WHERE username = %s
    """, (username,))

    user = cursor.fetchone()
    conn.close()

    if not user:
        return None

    stored_hash = user["password"].encode()

    if bcrypt.checkpw(password.encode("utf-8"), stored_hash):
        return {
            "id": user["id"],
            "username": user["username"],
            "role": user["role"]
        }

    return None


def get_users(search_text: str = ""):
    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT id, username, role, created_at
        FROM users
        WHERE 1=1
    """
    params = []

    if search_text:
        like_value = f"%{search_text}%"
        query += " AND (username LIKE %s OR role LIKE %s)"
        params.extend([like_value, like_value])

    query += " ORDER BY id DESC"

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    return rows


def get_user_by_id(user_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, username, password, role
        FROM users
        WHERE id = %s
    """, (user_id,))
    row = cursor.fetchone()
    conn.close()
    return row


def add_user(username: str, password: str, role: str):
    conn = get_connection()
    cursor = conn.cursor()

    hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode()

    cursor.execute("""
        INSERT INTO users (username, password, role)
        VALUES (%s, %s, %s)
    """, (username, hashed, role))

    conn.commit()
    conn.close()

def update_user(user_id: int, username: str, password: str, role: str):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE users
        SET username = %s, password = %s, role = %s
        WHERE id = %s
    """, (username, password, role, user_id))
    conn.commit()
    conn.close()


def delete_user(user_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
    conn.commit()
    conn.close()


# ---------------- PRODUCTS ----------------

def get_products(search_text: str = "", category_filter: str = "All"):
    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT id, name, brand, model, specifications, product_condition, category, price, stock_qty, status, track_serials, serial_numbers, created_at
        FROM products
        WHERE 1=1
    """
    params = []

    if search_text:
        like_value = f"%{search_text}%"
        query += " AND (name LIKE %s OR category LIKE %s)"
        params.extend([like_value, like_value])

    if category_filter and category_filter != "All":
        query += " AND category = %s"
        params.append(category_filter)

    query += " ORDER BY id DESC"

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()

    products = []
    for row in rows:
        serials_raw = row.get("serial_numbers")
        if isinstance(serials_raw, str):
            try:
                serials = json.loads(serials_raw)
            except Exception:
                serials = []
        elif isinstance(serials_raw, list):
            serials = serials_raw
        else:
            serials = []

        products.append({
            "id": row["id"],
            "name": row["name"],
            "brand": row.get("brand") or "",
            "model": row.get("model") or "",
            "specifications": row.get("specifications") or "",
            "product_condition": row.get("product_condition") or "",
            "category": row["category"],
            "price": float(row["price"]),
            "stock_qty": int(row["stock_qty"]),
            "status": row["status"],
            "track_serials": bool(row["track_serials"]),
            "serial_numbers": serials,
            "created_at": row["created_at"],
        })

    return products

def get_all_products():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, name, category, price, stock_qty, status
        FROM products
        ORDER BY name ASC
    """)
    rows = cursor.fetchall()
    conn.close()
    return rows


def get_product_by_id(product_id: int):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, name, brand, model, specifications, product_condition,
               category, price, stock_qty, status, track_serials, serial_numbers, created_at
        FROM products
        WHERE id = %s
    """, (product_id,))

    row = cursor.fetchone()
    conn.close()

    if not row:
        return None

    serials_raw = row.get("serial_numbers")
    if isinstance(serials_raw, str):
        try:
            serials = json.loads(serials_raw)
        except Exception:
            serials = []
    elif isinstance(serials_raw, list):
        serials = serials_raw
    else:
        serials = []

    return {
        "id": row["id"],
        "name": row["name"],
        "brand": row.get("brand") or "",
        "model": row.get("model") or "",
        "specifications": row.get("specifications") or "",
        "product_condition": row.get("product_condition") or "",
        "category": row["category"],
        "price": float(row["price"]),
        "stock_qty": int(row["stock_qty"]),
        "status": row["status"],
        "track_serials": bool(row["track_serials"]),
        "serial_numbers": serials,
        "created_at": row["created_at"],
    }

def add_product(
    name,
    brand,
    model,
    specifications,
    product_condition,
    category,
    price,
    stock_qty,
    track_serials=False,
    serial_numbers=None
):
    if serial_numbers is None:
        serial_numbers = []

    status = calculate_status(stock_qty)
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO products (
            name, brand, model, specifications, product_condition,
            category, price, stock_qty, status, track_serials, serial_numbers
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        name,
        brand,
        model,
        specifications,
        product_condition,
        category,
        price,
        stock_qty,
        status,
        1 if track_serials else 0,
        json.dumps(serial_numbers),
    ))
    conn.commit()
    conn.close()


def update_product(
    product_id: int,
    name: str,
    brand: str,
    model: str,
    specifications: str,
    product_condition: str,
    category: str,
    price: float,
    stock_qty: int
):
    status = calculate_status(stock_qty)
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE products
        SET name = %s,
            brand = %s,
            model = %s,
            specifications = %s,
            product_condition = %s,
            category = %s,
            price = %s,
            stock_qty = %s,
            status = %s
        WHERE id = %s
    """, (
        name,
        brand,
        model,
        specifications,
        product_condition,
        category,
        price,
        stock_qty,
        status,
        product_id
    ))
    conn.commit()
    conn.close()


def delete_product(product_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM products WHERE id = %s", (product_id,))
    conn.commit()
    conn.close()


def get_product_categories():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT category
        FROM products
        WHERE category IS NOT NULL AND TRIM(category) != ''
        ORDER BY category ASC
    """)
    rows = cursor.fetchall()
    conn.close()
    return [row["category"] for row in rows]


def get_product_serial_statuses(product_id: int):
    product = get_product_by_id(product_id)
    if not product:
        return []

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT serial_number
        FROM invoice_items
        WHERE product_id = %s AND TRIM(serial_number) != ''
        ORDER BY serial_number ASC
    """, (product_id,))
    sold_rows = cursor.fetchall()
    conn.close()

    available_serials = product.get("serial_numbers", [])
    sold_serials = [row["serial_number"] for row in sold_rows]

    all_serials = []

    for serial in sorted(available_serials):
        all_serials.append({
            "serial_number": serial,
            "status": "Available"
        })

    for serial in sorted(sold_serials):
        if serial not in available_serials:
            all_serials.append({
                "serial_number": serial,
                "status": "Sold"
            })

    return all_serials


def find_product_by_serial(serial_number: str):
    serial_number = serial_number.strip()
    if not serial_number:
        return None

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, name, category, price, stock_qty, status, track_serials, serial_numbers, created_at
        FROM products
        WHERE track_serials = 1
    """)
    rows = cursor.fetchall()
    conn.close()

    for row in rows:
        serials_raw = row.get("serial_numbers")
        if isinstance(serials_raw, str):
            try:
                serials = json.loads(serials_raw)
            except Exception:
                serials = []
        elif isinstance(serials_raw, list):
            serials = serials_raw
        else:
            serials = []

        if serial_number in serials:
            return {
                "id": row["id"],
                "name": row["name"],
                "category": row["category"],
                "price": float(row["price"]),
                "stock_qty": int(row["stock_qty"]),
                "status": row["status"],
                "track_serials": bool(row["track_serials"]),
                "serial_numbers": serials,
                "created_at": row["created_at"],
                "matched_serial": serial_number,
            }

    return None


# ---------------- CUSTOMERS ----------------

def get_customers(search_text: str = ""):
    conn = get_connection()
    cursor = conn.cursor()

    query = """
        SELECT id, name, phone, email, address
        FROM customers
        WHERE 1=1
    """
    params = []

    if search_text:
        like_value = f"%{search_text}%"
        query += " AND (name LIKE %s OR phone LIKE %s OR email LIKE %s)"
        params.extend([like_value, like_value, like_value])

    query += " ORDER BY id DESC"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    return rows


def get_all_customers():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, name, phone, email, address
        FROM customers
        ORDER BY name ASC
    """)
    rows = cursor.fetchall()
    conn.close()
    return rows


def add_customer(name: str, phone: str, email: str, address: str):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO customers (name, phone, email, address)
        VALUES (%s, %s, %s, %s)
    """, (name, phone, email, address))
    conn.commit()
    conn.close()


def update_customer(customer_id: int, name: str, phone: str, email: str, address: str):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE customers
        SET name = %s, phone = %s, email = %s, address = %s
        WHERE id = %s
    """, (name, phone, email, address, customer_id))
    conn.commit()
    conn.close()


def delete_customer(customer_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM customers WHERE id = %s", (customer_id,))
    conn.commit()
    conn.close()


def get_customer_by_id(customer_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, name, phone, email, address
        FROM customers
        WHERE id = %s
    """, (customer_id,))
    row = cursor.fetchone()
    conn.close()
    return row


# ---------------- INVOICES ----------------

def generate_invoice_number():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) AS total FROM invoices")
    total = cursor.fetchone()["total"]
    conn.close()
    return f"INV-{int(total) + 1:05d}"


def get_invoices():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, invoice_number, customer_name, total_amount, payment_status, created_at
        FROM invoices
        ORDER BY id DESC
    """)
    rows = cursor.fetchall()
    conn.close()
    return rows


def get_invoice_by_id(invoice_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, invoice_number, customer_id, customer_name, total_amount, payment_status, created_at
        FROM invoices
        WHERE id = %s
    """, (invoice_id,))
    row = cursor.fetchone()
    conn.close()
    return row


def get_invoice_items(invoice_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT product_name, price, quantity, serial_number, line_total
        FROM invoice_items
        WHERE invoice_id = %s
        ORDER BY id ASC
    """, (invoice_id,))
    rows = cursor.fetchall()
    conn.close()
    return rows


def mark_invoice_as_paid(invoice_id: int):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE invoices
        SET payment_status = 'PAID'
        WHERE id = %s
    """, (invoice_id,))
    conn.commit()
    conn.close()


def create_invoice(customer_id: int, items: list):
    if not items:
        raise ValueError("Invoice must contain at least one item.")

    conn = get_connection()
    cursor = conn.cursor()

    try:
        customer = get_customer_by_id(customer_id)
        if not customer:
            raise ValueError("Customer not found.")

        total_amount = 0
        prepared_items = []

        for item in items:
            product = get_product_by_id(item["product_id"])
            if not product:
                raise ValueError(f"Product not found: {item['product_name']}")

            quantity = int(item["quantity"])
            if quantity <= 0:
                raise ValueError(f"Invalid quantity for {product['name']}")

            serial_number = item.get("serial_number", "").strip()

            if product.get("track_serials", False):
                if quantity != 1:
                    raise ValueError(f"Serialized product {product['name']} must have quantity 1 per line.")

                available_serials = product.get("serial_numbers", [])
                if serial_number == "":
                    raise ValueError(f"Serial number is required for {product['name']}")

                if serial_number not in available_serials:
                    raise ValueError(f"Serial number {serial_number} is not available for {product['name']}.")

                new_serials = [s for s in available_serials if s != serial_number]
                new_stock = len(new_serials)
            else:
                if product["stock_qty"] < quantity:
                    raise ValueError(
                        f"Not enough stock for {product['name']}. Available: {product['stock_qty']}"
                    )

                new_stock = product["stock_qty"] - quantity
                new_serials = product.get("serial_numbers", [])

            price = float(product["price"])
            line_total = price * quantity
            total_amount += line_total

            prepared_items.append({
                "product_id": product["id"],
                "product_name": product["name"],
                "price": price,
                "quantity": quantity,
                "serial_number": serial_number,
                "line_total": line_total,
                "new_stock": new_stock,
                "new_serials": new_serials,
            })

        invoice_number = generate_invoice_number()

        cursor.execute("""
            INSERT INTO invoices (invoice_number, customer_id, customer_name, total_amount, payment_status)
            VALUES (%s, %s, %s, %s, %s)
        """, (invoice_number, customer_id, customer["name"], total_amount, "UNPAID"))

        invoice_id = cursor.lastrowid

        for item in prepared_items:
            cursor.execute("""
                INSERT INTO invoice_items (
                    invoice_id, product_id, product_name, price, quantity, serial_number, line_total
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                invoice_id,
                item["product_id"],
                item["product_name"],
                item["price"],
                item["quantity"],
                item["serial_number"],
                item["line_total"],
            ))

            cursor.execute("""
                UPDATE products
                SET stock_qty = %s, status = %s, serial_numbers = %s
                WHERE id = %s
            """, (
                item["new_stock"],
                calculate_status(item["new_stock"]),
                json.dumps(item["new_serials"]),
                item["product_id"]
            ))

        conn.commit()
        return invoice_id, invoice_number

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def find_serial_usage(serial_number: str):
    serial_number = serial_number.strip()
    if not serial_number:
        return None

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            ii.serial_number,
            ii.product_name,
            i.invoice_number,
            i.customer_name,
            i.created_at
        FROM invoice_items ii
        JOIN invoices i ON ii.invoice_id = i.id
        WHERE ii.serial_number = %s
        ORDER BY ii.id DESC
        LIMIT 1
    """, (serial_number,))
    sold_row = cursor.fetchone()

    if sold_row:
        conn.close()
        return {
            "serial_number": sold_row["serial_number"],
            "product_name": sold_row["product_name"],
            "status": "Sold",
            "invoice_number": sold_row["invoice_number"],
            "customer_name": sold_row["customer_name"],
            "created_at": sold_row["created_at"],
        }

    cursor.execute("""
        SELECT id, name, serial_numbers
        FROM products
        WHERE track_serials = 1
    """)
    rows = cursor.fetchall()
    conn.close()

    for row in rows:
        serials_raw = row.get("serial_numbers")
        if isinstance(serials_raw, str):
            try:
                serials = json.loads(serials_raw)
            except Exception:
                serials = []
        elif isinstance(serials_raw, list):
            serials = serials_raw
        else:
            serials = []

        if serial_number in serials:
            return {
                "serial_number": serial_number,
                "product_name": row["name"],
                "status": "In Stock",
                "invoice_number": "",
                "customer_name": "",
                "created_at": "",
            }

    return None


# ---------------- DASHBOARD ----------------

def get_total_products():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) AS total FROM products")
    total = cursor.fetchone()["total"]
    conn.close()
    return total


def get_low_stock_products_count():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT COUNT(*) AS total
        FROM products
        WHERE stock_qty > 0 AND stock_qty <= 5
    """)
    total = cursor.fetchone()["total"]
    conn.close()
    return total


def get_total_customers():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) AS total FROM customers")
    total = cursor.fetchone()["total"]
    conn.close()
    return total


def get_total_invoices():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) AS total FROM invoices")
    total = cursor.fetchone()["total"]
    conn.close()
    return total


def get_total_sales():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COALESCE(SUM(total_amount), 0) AS total FROM invoices")
    total = cursor.fetchone()["total"]
    conn.close()
    return float(total)


def get_recent_invoices(limit=5):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT invoice_number, customer_name, total_amount, created_at
        FROM invoices
        ORDER BY id DESC
        LIMIT %s
    """, (limit,))
    rows = cursor.fetchall()
    conn.close()
    return rows


def get_low_stock_products(limit=5):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT name, category, stock_qty, status
        FROM products
        WHERE stock_qty > 0 AND stock_qty <= 5
        ORDER BY stock_qty ASC, name ASC
        LIMIT %s
    """, (limit,))
    rows = cursor.fetchall()
    conn.close()
    return rows

def get_sales_report_summary():
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            COUNT(*) AS total_invoices,
            COALESCE(SUM(total_amount), 0) AS total_sales,
            COALESCE(SUM(CASE WHEN payment_status = 'PAID' THEN total_amount ELSE 0 END), 0) AS paid_sales,
            COALESCE(SUM(CASE WHEN payment_status = 'UNPAID' THEN total_amount ELSE 0 END), 0) AS unpaid_sales
        FROM invoices
    """)
    row = cursor.fetchone()
    conn.close()
    return row


def get_inventory_report_summary():
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            COUNT(*) AS total_products,
            COALESCE(SUM(stock_qty), 0) AS total_stock_units,
            COALESCE(SUM(price * stock_qty), 0) AS stock_value
        FROM products
    """)
    row = cursor.fetchone()
    conn.close()
    return row


def get_low_stock_report():
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, name, category, stock_qty, status, price
        FROM products
        WHERE stock_qty > 0 AND stock_qty <= 5
        ORDER BY stock_qty ASC, name ASC
    """)
    rows = cursor.fetchall()
    conn.close()
    return rows


def get_recent_sales_report(limit=10):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT invoice_number, customer_name, total_amount, payment_status, created_at
        FROM invoices
        ORDER BY id DESC
        LIMIT %s
    """, (limit,))
    rows = cursor.fetchall()
    conn.close()
    return rows

def export_sales_report_to_excel():
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT invoice_number, customer_name, total_amount, payment_status, created_at
        FROM invoices
        ORDER BY id DESC
    """)
    rows = cursor.fetchall()
    conn.close()

    reports_dir = Path("reports")
    reports_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = reports_dir / f"sales_report_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = ["Invoice Number", "Customer", "Total Amount", "Payment Status", "Created At"]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="10B981")
    header_font = Font(color="FFFFFF", bold=True)

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_sales = 0.0

    for row in rows:
        total_amount = float(row["total_amount"])
        total_sales += total_amount

        ws.append([
            row["invoice_number"],
            row["customer_name"],
            total_amount,
            row["payment_status"],
            str(row["created_at"]),
        ])

    # Currency format for total amount column
    for row_num in range(2, ws.max_row + 1):
        ws.cell(row=row_num, column=3).number_format = '₦#,##0.00'

    # Summary rows
    summary_start = ws.max_row + 2
    ws.cell(row=summary_start, column=1, value="Total Invoices")
    ws.cell(row=summary_start, column=2, value=len(rows))

    ws.cell(row=summary_start + 1, column=1, value="Total Sales")
    total_sales_cell = ws.cell(row=summary_start + 1, column=2, value=total_sales)
    total_sales_cell.number_format = '₦#,##0.00'

    # Bold summary
    for r in [summary_start, summary_start + 1]:
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2).font = Font(bold=True)

    # Auto width
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value_length = len(str(cell.value)) if cell.value is not None else 0
                if value_length > max_length:
                    max_length = value_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = max_length + 4

    wb.save(file_path)
    return str(file_path)